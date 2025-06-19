from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

thin = Side(border_style = "thin", color = "000000")
none = Side(border_style = 'none')

cells_profiles = {
    0: {
        'alignment': Alignment(wrap_text = True, horizontal = 'left', vertical = 'center'),
        'font': Font(size = 7.5),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    1: {
        'alignment': Alignment(horizontal = 'right', vertical = 'center'),
        'font': Font(size = 10, bold = True),
        'border': Border(left = thin, top = thin, right = none, bottom = thin),
        'fill': PatternFill(fill_type = 'solid', fgColor = '999999')
    },

    2: {
        'alignment': Alignment(horizontal = 'right', vertical = 'center'),
        'font': Font(size = 9),
        'border': Border(left = none, top = thin, right = thin, bottom = thin),
        'fill': PatternFill(fill_type = 'solid', fgColor = '999999')
    },

    3: {
        'alignment': Alignment(horizontal = 'left', vertical = 'center'),
        'font': Font(size = 9, bold = True),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    4: {
        'alignment': Alignment(wrap_text = True, horizontal = 'left', vertical = 'center'),
        'font': Font(size = 9),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    5: {
        'alignment': Alignment(horizontal = 'center', vertical = 'center'),
        'font': Font(size = 9, bold = True),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill(fill_type = 'solid', fgColor = '999999')
    },

    6: {
        'alignment': Alignment(horizontal = 'left', vertical = 'center'),
        'font': Font(size = 9, bold = True),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill(fill_type = 'solid', fgColor = 'DDDDDD')
    },

    7: {
        'alignment': Alignment(horizontal = 'center', vertical = 'center'),
        'font': Font(size = 9, bold = True),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill(fill_type = 'solid', fgColor = 'DDDDDD')
    },

    8: {
        'alignment': Alignment(horizontal = 'center', vertical = 'center'),
        'font': Font(size = 9),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    9: {
        'alignment': Alignment(wrap_text = True, horizontal = 'center', vertical = 'center'),
        'font': Font(size = 9),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    10: {
        'alignment': Alignment(horizontal = 'left', vertical = 'center'),
        'font': Font(size = 9),
        'border': Border(left = thin, top = thin, right = none, bottom = thin),
        'fill': PatternFill()
    },

    11: {
        'alignment': Alignment(horizontal = 'left', vertical = 'center'),
        'font': Font(size = 9),
        'border': Border(left = none, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    12: {
        'alignment': Alignment(horizontal = 'center', vertical = 'center'),
        'font': Font(size = 9),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    13: {
        'alignment': Alignment(),
        'font': Font(),
        'border': Border(left = thin, top = thin, right = thin, bottom = thin),
        'fill': PatternFill()
    },

    14: {
        'alignment': Alignment(),
        'font': Font(),
        'border': Border(left = none, top = none, right = none, bottom = none),
        'fill': PatternFill()
    }
}

def xlsx_AQTW_template_creator(filename, destination_directory):

    type_aqtw = filename.split('_')[2].upper()

    cells_layout = []
    cells_layout.append({'ROW_HEIGHT': [15], 'B-I': [f'REPORT AUTOMATICO COLLAUDO {type_aqtw} MEVO', 1], 'J-N': ['Mod. PROD. 03 Rev. 0 del 05/06/2025', 2]})
    cells_layout.append({'ROW_HEIGHT': [50], 'B-G': ['LOGO', 13], 'H-N': ['G.S.I. srl\nVia dell\'Artigianato 44\n31047 Ponte di Piave (TV)\nTel 0422 289828 .Fax. 0422 759905', 0]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Matricola', 3], 'D-G': ['', 12], 'H': ['Data', 3], 'I-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Commessa', 3], 'D-G': ['', 12], 'H': ['Conferma', 3], 'I-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [15], 'B-N': ['DATI GENERALI', 5]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Versione software', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Versione OS', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [20], 'B-F': ['POMPE A BORDO', 6], 'G-J': ['108 - 112', 7], 'K-N': ['116 - 120', 7]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Pompa IMPIANTO', 4], 'G-J': ['PARA R 15-130 / 9-87 / IPWM2', 8], 'K-N': ['PARA MAXO R 25-180-10-F22 GSY', 8]})
    cells_layout.append({'ROW_HEIGHT': [20], 'B-F': ['SONDE E TRASDUTTORI', 6], 'G-N': ['LETTURA', 7]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura Mandata Impianto', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura Ritorno Impianto', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura ACS', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Alta Pressione', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Bassa Pressione', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-N': ['INVERTER COMPRESSORE', 6]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Velocità Compressore', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [20], 'B-F': ['Verifiche', 6], 'G-N': ['Verificato', 7]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica EEV 1 Movimento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica EEV 2 Movimento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Resistenza Vaschetta Funzionamento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Bobina 4 VIE Funzionamento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Sonde NTC macchina', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica TRASDUTTORI', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Pompa Geotermica', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Pompa Mix', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Valvola Mix', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Gas Refrigerante', 4], 'D-F': ['R290', 9], 'G-N': ['Qta [kg]:', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-N': ['VENTILATORI', 6]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Verifica Ventilatore 1 (alto)', 4], 'D-F': ['Modello:VHE01195338', 9], 'G-N': ['Comando 0..10Vdc:', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Verifica Ventilatore 2 (basso)', 4], 'D-F': ['Modello:VHE01195338', 9], 'G-N': ['Comando 0..10Vdc:', 12]})

    return create_template(filename, destination_directory, cells_layout)

def xlsx_MEVO_template_creator(filename, destination_directory, is_pump_acs_installed):

    type_mevo = filename.split('_')[2].upper()

    cells_layout = []
    cells_layout.append({'ROW_HEIGHT': [15], 'B-I': [f'REPORT AUTOMATICO COLLAUDO {type_mevo} MEVO', 1], 'J-N': ['Mod. PROD. 03 Rev. 0 del 05/06/2025', 2]})
    cells_layout.append({'ROW_HEIGHT': [50], 'B-G': ['LOGO', 13], 'H-N': ['G.S.I. srl\nVia dell\'Artigianato 44\n31047 Ponte di Piave (TV)\nTel 0422 289828 .Fax. 0422 759905', 0]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Matricola', 3], 'D-G': ['', 12], 'H': ['Data', 3], 'I-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Commessa', 3], 'D-G': ['', 12], 'H': ['Conferma', 3], 'I-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [15], 'B-N': ['DATI GENERALI', 5]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Versione software', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Versione OS', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [20], 'B-F': ['POMPE A BORDO', 6], 'G-J': ['108 - 112', 7], 'K-N': ['116 - 120', 7]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Pompa IMPIANTO', 4], 'G-J': ['PARA R 15-130 / 9-87 / IPWM2', 8], 'K-N': ['PARA MAXO R 25-180-10-F22 GSY', 8]})
    if is_pump_acs_installed:
        cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Pompa ACS', 4], 'G-J': ['PARA R 15-130 / 9-87 / IPWM2', 8], 'K-N': ['PARA MAXO R 25-180-10-F22 GSY', 8]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Portata IMPIANTO', 4], 'G-N': ['', 12]})
    if is_pump_acs_installed:
        cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Portata ACS', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [20], 'B-F': ['SONDE E TRASDUTTORI', 6], 'G-N': ['LETTURA', 7]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura Mandata Impianto', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura Ritorno Impianto', 4], 'G-N': ['', 12]})
    # cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura Mandata Geotermico', 4], 'G-N': ['', 12]})
    # cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura Ritorno Geotermico', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Temperatura ACS', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Alta Pressione', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Bassa Pressione', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-N': ['INVERTER COMPRESSORE', 6]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Velocità Compressore', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [20], 'B-F': ['Verifiche', 6], 'G-N': ['Verificato', 7]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica EEV 1 Movimento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica EEV 2 Movimento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Resistenza Vaschetta Funzionamento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Bobina 4 VIE Funzionamento', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica Sonde NTC macchina', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-F': ['Verifica TRASDUTTORI', 4], 'G-N': ['', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Gas Refrigerante', 4], 'D-F': ['R290', 9], 'G-N': ['Qta [kg]:', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-N': ['VENTILATORI', 6]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Verifica Ventilatore 1 (alto)', 4], 'D-F': ['Modello:VHE01195338', 9], 'G-N': ['Comando 0..10Vdc:', 12]})
    cells_layout.append({'ROW_HEIGHT': [13], 'B-C': ['Verifica Ventilatore 2 (basso)', 4], 'D-F': ['Modello:VHE01195338', 9], 'G-N': ['Comando 0..10Vdc:', 12]})

    return create_template(filename, destination_directory, cells_layout)


def create_template(filename, destination_directory, cells_layout):
    wb = Workbook()
    ws = wb.active

    start_row = 2
    start_col = 'B'

    end_row = len(cells_layout) + start_row - 1
    end_col = 'N'

    for row in range(0, len(cells_layout)):
        cell_row = row + start_row

        for key, value in cells_layout[row].items():

            if key == 'ROW_HEIGHT':
                ws.row_dimensions[cell_row].height = value[0]
            else:
                first_merged_cell_col = key.split('-')[0] if '-' in key else key
                last_merged_cell_col = key.split('-')[1] if '-' in key else key

                ws.merge_cells(f'{first_merged_cell_col}{cell_row}:{last_merged_cell_col}{cell_row}')

                cell_address = f'{first_merged_cell_col}{cell_row}'
                if value[0] != 'LOGO':
                    ws[cell_address] = value[0]

                    ws[cell_address].alignment = cells_profiles[value[1]]['alignment']
                    ws[cell_address].font = cells_profiles[value[1]]['font']
                    ws[cell_address].fill = cells_profiles[value[1]]['fill']

                else:
                    def excel_col_width_to_pixels(width):
                        return int(width * 7)

                    def excel_row_height_to_pixels(height):
                        return int(height * 1.33)

                    ws.column_dimensions[first_merged_cell_col].width = 20
                    ws.row_dimensions[cell_row].height = 50

                    col_width = excel_col_width_to_pixels(ws.column_dimensions[first_merged_cell_col].width)
                    row_height = excel_row_height_to_pixels(ws.row_dimensions[cell_row].height)

                    img_path = 'src\\resources\\gsi_logo.png'
                    img_pil = PILImage.open(img_path)

                    logo_w, logo_h = img_pil.size
                    scale_w = col_width / logo_w
                    scale_h = row_height / logo_h
                    scale = min(scale_w, scale_h) * 1.4

                    new_w = int(logo_w * scale)
                    new_h = int(logo_h * scale)

                    img_pil = img_pil.resize((new_w, new_h), PILImage.LANCZOS)

                    final_img = PILImage.new("RGBA", (col_width, row_height), (255, 255, 255, 0))
                    left = (col_width - new_w) // 2
                    top = (row_height - new_h) // 2
                    final_img.paste(img_pil, (left, top), img_pil if img_pil.mode == "RGBA" else None)

                    temp_logo_path = "src\\resources\\logo_resized.png"
                    final_img.save(temp_logo_path)
                    img = Image(temp_logo_path)
                    ws.add_image(img, f'{first_merged_cell_col}{cell_row}')

                for merged_cell_col in range(ord(first_merged_cell_col), ord(last_merged_cell_col) + 1):
                    cell_address = f'{chr(merged_cell_col)}{cell_row}'
                    ws[cell_address].border = cells_profiles[value[1]]['border']
                
    wb.save(f'{destination_directory}\\{filename}.xlsx')

    return f'{destination_directory}\\{filename}.xlsx'   