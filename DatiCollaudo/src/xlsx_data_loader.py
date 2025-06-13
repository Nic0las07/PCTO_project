from pathlib import Path
from xlsx_layout_creator import xlsx_template_creator
from openpyxl import load_workbook
import re

template_start_row = 2
template_end_row = 61
template_start_col = 'B'
template_end_col = 'N'

directory = Path('formatted_csv')

final_directory = Path('final_data_files')
final_directory.mkdir(exist_ok = True)

for current_file in directory.iterdir():

    if(current_file.suffix == '.csv'):
        print(f'\n\nPreparazione del file {current_file.name} in corso......')
        current_filename = current_file.name.removesuffix('.csv')
        xlsx_template = open(xlsx_template_creator(current_filename, final_directory.name))
        current_csv = current_file.open()

        wb = load_workbook(xlsx_template.name)
        ws = wb.active
        merged_cells_ranges = {}

        for merged_range in ws.merged_cells.ranges:
            merged_range_parts = str(merged_range).split(':')
            if len(merged_range_parts) < 2 : merged_range_parts.append(merged_range_parts[0])
            merged_cells_ranges[merged_range_parts[0]] = merged_range_parts[1]

        #Raccolta dati

        csv_data = {}
        info_parts = current_filename.split('_')
        csv_data['Matricola'] = info_parts[1]
        csv_data['Commessa'] = info_parts[0]
        csv_data['Conferma'] = info_parts[1][:5]

        for line in current_csv:
            line_parts = line.split(';')
            if line_parts[0].lower() == 'time':
                line_parts[0] = 'Data'
                line_parts[1] = line_parts[1][:line_parts[1].find('T')]

            csv_data[line_parts[0].replace('\"', '')] = line_parts[1] + ' ' + line_parts[2]

        # Gestisce e unisce i campi delle versioni presenti nei dati CSV,
        # raggruppando e ordinando i valori delle versioni in un'unica stringa per ogni tipo di versione.

        versions = {}

        for key, value in csv_data.items():
            if key.find('Versione') != -1:
                version_order = ''.join(ch if ch.isdigit() else '' for ch in key)
                versions.setdefault(key.replace(version_order, ''), {})[version_order] = value

        keys_to_delete = [key for key in csv_data if 'Version' in key]

        for key in keys_to_delete : csv_data.pop(key, None)

        for key, value in versions.items():
            version_size = max(int(size) for size in value)
            versions_correct_order = [None] * version_size
            version_num = ''

            for version_order, version_value in value.items():
                versions_correct_order[int(version_order) - 1] = version_value
            
            for num_version in versions_correct_order:
                if num_version : version_num += num_version  + '.'

            version_num = version_num[:-1]

            csv_data[key] = version_num

        percentage_to_add = 100 / len(csv_data)
        complete_percentage = 0

        #Scrittura automatica dei dati nelle rispettive celle
        can_put_data = True
        
        for key, value in csv_data.items():
            complete_percentage += percentage_to_add
            print(f"Lavoro al {complete_percentage:.2f}%")

            key_name = key
            key_name = ''.join(ch if ch.isalnum() else ' ' for ch in key_name)
            
            for index in range(0, len(key_name)):
                if key_name[index].isupper() and index != 0 and key_name[index - 1].islower():
                    key_name = key_name[:index] + ' ' + key_name[index:]
            
            key_name = ''.join(ch.lower() if ch.isalpha() else ch for ch in key_name)
            field_to_search = key_name.split(' ')

            for first_merged_col_address, last_merged_col_address in merged_cells_ranges.items():
                if not(ws[first_merged_col_address].value): continue
                if re.search(r':\s*$', ws[first_merged_col_address].value): continue
                if last_merged_col_address[0] == template_end_col: continue
                can_put_data = True

                for field_part in field_to_search:
                    if not(re.search(rf'\b{field_part}', ws[first_merged_col_address].value.lower())):
                        can_put_data = False
                        break
                if not(can_put_data): continue

                next_merged_cell_address = chr(ord(last_merged_col_address[0]) + 1) + last_merged_col_address[1:]
                while ord(next_merged_cell_address[0]) < ord(template_end_col):
                    if not(ws[next_merged_cell_address].value):
                        ws[next_merged_cell_address] = value
                        break
                    elif re.search(r':\s*$', ws[next_merged_cell_address].value):
                        ws[next_merged_cell_address] = ws[next_merged_cell_address].value + ' ' + value
                        break
                    next_merged_cell_col = chr(ord(merged_cells_ranges[next_merged_cell_address][0]) + 1)
                    next_merged_cell_row = str(merged_cells_ranges[next_merged_cell_address][1:])
                    next_merged_cell_address =  next_merged_cell_col + next_merged_cell_row

                

        print(f'Preparazione finita. Nuovo file {xlsx_template.name[xlsx_template.name.find('\\')+1:]} creato in {final_directory.name}\n\n')

        wb.save(xlsx_template.name)